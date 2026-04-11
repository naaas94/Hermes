"""File type detection and preflight classification."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from hermes.models import FileType, PreflightResult
from hermes.normalization.chunker import CHARS_PER_TOKEN

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
PDF_EXTENSIONS = {".pdf"}

# Magic bytes for common file types
MAGIC_XLSX = b"PK"  # ZIP archive (xlsx is a ZIP)
MAGIC_PDF = b"%PDF"

# If a PDF page yields fewer than this many characters, it's likely scanned
TEXT_LAYER_CHAR_THRESHOLD = 50

# Excel token estimate: full scan when summed sheet dimensions (max_row per sheet) stay
# below this; otherwise prefix-sample each sheet and extrapolate from dimensions.
EXCEL_PREFLIGHT_FULL_SCAN_MAX_ROWS = 10_000
EXCEL_PREFLIGHT_PREFIX_ROWS_PER_SHEET = 500


def _excel_row_chars(row: tuple[Any, ...]) -> int:
    return sum(len(str(c)) for c in row if c is not None)


def _excel_total_dim_rows(wb: Any) -> int:
    total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        mr = ws.max_row
        total += max(mr, 1) if mr is not None and mr >= 1 else 1
    return total


def _excel_full_scan_chars(wb: Any) -> int:
    total_chars = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            total_chars += _excel_row_chars(row)
    return total_chars


def _excel_sampled_chars(wb: Any, prefix_rows: int) -> int:
    total_chars = 0
    for name in wb.sheetnames:
        ws = wb[name]
        mr = ws.max_row
        if mr is not None and mr >= 1:
            effective_max = mr
        else:
            effective_max = 1
        cap = min(prefix_rows, effective_max)
        sample_chars = 0
        sample_rows = 0
        for row in ws.iter_rows(values_only=True, max_row=cap):
            sample_chars += _excel_row_chars(row)
            sample_rows += 1
        if sample_rows <= 0:
            continue
        total_chars += int((sample_chars / sample_rows) * effective_max)
    return total_chars


def _excel_estimated_chars(wb: Any) -> int:
    if _excel_total_dim_rows(wb) <= EXCEL_PREFLIGHT_FULL_SCAN_MAX_ROWS:
        return _excel_full_scan_chars(wb)
    return _excel_sampled_chars(wb, EXCEL_PREFLIGHT_PREFIX_ROWS_PER_SHEET)


def detect_file_type(file_path: Path) -> FileType:
    """Detect whether a file is Excel, text-PDF, scanned-PDF, or unknown."""
    ext = file_path.suffix.lower()

    if ext in EXCEL_EXTENSIONS:
        return FileType.EXCEL

    if ext in PDF_EXTENSIONS:
        return _classify_pdf(file_path)

    magic = _read_magic(file_path)
    if magic.startswith(MAGIC_XLSX):
        return FileType.EXCEL
    if magic.startswith(MAGIC_PDF):
        return _classify_pdf(file_path)

    return FileType.UNKNOWN


def _read_magic(file_path: Path, n: int = 8) -> bytes:
    try:
        with open(file_path, "rb") as f:
            return f.read(n)
    except OSError:
        return b""


def _classify_pdf(file_path: Path) -> FileType:
    """Determine if a PDF has a usable text layer or is scanned."""
    try:
        import pymupdf
    except ImportError:
        return FileType.PDF_TEXT  # assume text if pymupdf not available

    doc: Any = pymupdf.open(str(file_path))  # type: ignore[no-untyped-call]
    try:
        text_chars = 0
        pages_checked = min(len(doc), 5)  # sample first 5 pages
        for i in range(pages_checked):
            page: Any = doc[i]
            text = page.get_text("text")
            text_chars += len(text.strip())
            del page

        avg_chars = text_chars / max(pages_checked, 1)
        if avg_chars < TEXT_LAYER_CHAR_THRESHOLD:
            return FileType.PDF_SCANNED
        return FileType.PDF_TEXT
    finally:
        doc.close()


def run_preflight(file_path: Path) -> PreflightResult:
    """Run full preflight classification on a file."""
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    file_type = detect_file_type(file_path)
    page_count = 0
    has_text_layer = False
    estimated_tokens = 0

    if file_type in (FileType.PDF_TEXT, FileType.PDF_SCANNED):
        try:
            import pymupdf

            doc: Any = pymupdf.open(str(file_path))  # type: ignore[no-untyped-call]
            page_count = len(doc)
            has_text_layer = file_type == FileType.PDF_TEXT

            if has_text_layer:
                total_chars = 0
                for i in range(len(doc)):
                    page = doc[i]
                    total_chars += len(page.get_text("text"))
                    del page
                estimated_tokens = total_chars // CHARS_PER_TOKEN
            doc.close()
        except ImportError:
            pass

    elif file_type == FileType.EXCEL:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            page_count = len(wb.sheetnames)
            total_chars = _excel_estimated_chars(wb)
            estimated_tokens = total_chars // CHARS_PER_TOKEN
            wb.close()
        except ImportError:
            pass

    return PreflightResult(
        file_type=file_type,
        page_count=page_count,
        has_text_layer=has_text_layer,
        estimated_tokens=estimated_tokens,
        file_name=file_path.name,
        file_size_bytes=file_path.stat().st_size,
    )
