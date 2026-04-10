"""Excel -> Markdown normalizer using openpyxl in read_only streaming mode."""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path
from typing import TYPE_CHECKING

from hermes.ingestion.storage import get_normalized_dir
from hermes.models import FileType, NormalizedPage

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

ROWS_PER_CHUNK = 50


def normalize_excel(
    file_path: Path,
    job_id: str,
    page_indices: frozenset[int] | None = None,
    on_page_done: Callable[[int], None] | None = None,
) -> list[NormalizedPage]:
    """Convert an Excel file to per-sheet Markdown files on disk.

    Uses openpyxl read_only mode to avoid loading the entire workbook into RAM.
    Rows are processed in chunks of ROWS_PER_CHUNK and flushed to disk immediately.
    """
    import openpyxl

    out_dir = get_normalized_dir(job_id)
    pages: list[NormalizedPage] = []

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            if page_indices is not None and sheet_idx not in page_indices:
                continue
            ws = wb[sheet_name]
            md_path = out_dir / f"sheet_{sheet_idx}.md"
            char_count = _write_sheet_markdown(ws, sheet_name, md_path)
            pages.append(
                NormalizedPage(
                    page_index=sheet_idx,
                    markdown_path=md_path,
                    source_type=FileType.EXCEL,
                    char_count=char_count,
                )
            )
            if on_page_done is not None:
                on_page_done(sheet_idx)
    finally:
        wb.close()

    return pages


def _write_sheet_markdown(ws: Worksheet, sheet_name: str, md_path: Path) -> int:
    """Stream rows from a worksheet into a Markdown file, chunk by chunk."""
    char_count = 0
    header_written = False
    headers: list[str] = []
    row_buffer: list[list[str]] = []

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# {sheet_name}\n\n")
        char_count += len(sheet_name) + 4

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]

            if not header_written:
                headers = cells
                header_written = True
                continue

            row_buffer.append(cells)

            if len(row_buffer) >= ROWS_PER_CHUNK:
                chunk_text = _format_markdown_table(headers, row_buffer)
                f.write(chunk_text)
                char_count += len(chunk_text)
                row_buffer = []

        if row_buffer:
            chunk_text = _format_markdown_table(headers, row_buffer)
            f.write(chunk_text)
            char_count += len(chunk_text)

        if not header_written:
            f.write("*Empty sheet*\n")
            char_count += 14

    return char_count


def _format_markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    """Format rows as a Markdown table with the given headers."""
    if not headers:
        return ""

    col_count = len(headers)
    lines: list[str] = []

    header_line = "| " + " | ".join(headers) + " |"
    sep_line = "| " + " | ".join("---" for _ in headers) + " |"
    lines.append(header_line)
    lines.append(sep_line)

    for row in rows:
        padded = row + [""] * (col_count - len(row)) if len(row) < col_count else row[:col_count]
        escaped = [c.replace("|", "\\|").replace("\n", " ") for c in padded]
        lines.append("| " + " | ".join(escaped) + " |")

    lines.append("")
    return "\n".join(lines) + "\n"
