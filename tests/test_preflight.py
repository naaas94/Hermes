"""Tests for preflight file classification."""

from __future__ import annotations

from pathlib import Path

import pytest

from hermes.models import FileType
from hermes.normalization.chunker import CHARS_PER_TOKEN


def test_detect_excel(sample_excel: Path):
    if not sample_excel.exists():
        pytest.skip("Run generate_fixtures.py first")
    from hermes.ingestion.preflight import detect_file_type
    assert detect_file_type(sample_excel) == FileType.EXCEL


def test_detect_pdf_text(sample_pdf_text: Path):
    if not sample_pdf_text.exists():
        pytest.skip("Run generate_fixtures.py first")
    from hermes.ingestion.preflight import detect_file_type
    assert detect_file_type(sample_pdf_text) == FileType.PDF_TEXT


def test_detect_unknown(tmp_path: Path):
    unknown = tmp_path / "file.txt"
    unknown.write_text("hello")
    from hermes.ingestion.preflight import detect_file_type
    assert detect_file_type(unknown) == FileType.UNKNOWN


def test_preflight_excel(sample_excel: Path):
    if not sample_excel.exists():
        pytest.skip("Run generate_fixtures.py first")
    pytest.importorskip("openpyxl")
    import openpyxl

    from hermes.ingestion import preflight as preflight_mod
    from hermes.ingestion.preflight import run_preflight

    wb = openpyxl.load_workbook(sample_excel, read_only=True, data_only=True)
    ref_chars = preflight_mod._excel_full_scan_chars(wb)
    wb.close()
    ref_tokens = ref_chars // CHARS_PER_TOKEN

    result = run_preflight(sample_excel)
    assert result.file_type == FileType.EXCEL
    assert result.page_count >= 1
    assert result.file_name == "sample.xlsx"
    assert result.estimated_tokens == ref_tokens


def test_preflight_pdf(sample_pdf_text: Path):
    if not sample_pdf_text.exists():
        pytest.skip("Run generate_fixtures.py first")
    from hermes.ingestion.preflight import run_preflight
    result = run_preflight(sample_pdf_text)
    assert result.file_type == FileType.PDF_TEXT
    assert result.page_count == 2
    assert result.has_text_layer is True


def test_preflight_missing_file():
    from hermes.ingestion.preflight import run_preflight
    with pytest.raises(FileNotFoundError):
        run_preflight(Path("/nonexistent/file.pdf"))


def test_preflight_excel_sampled_uniform_matches_full_scan(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Forced sampling (tiny full-scan threshold) should match full scan for uniform rows."""
    openpyxl = pytest.importorskip("openpyxl")
    import hermes.ingestion.preflight as preflight_mod
    from hermes.ingestion.preflight import run_preflight

    path = tmp_path / "uniform.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(5):
        ws.append(["aaaa"])
    wb.save(path)

    wb_ref = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ref_tokens = preflight_mod._excel_full_scan_chars(wb_ref) // CHARS_PER_TOKEN
    wb_ref.close()

    monkeypatch.setattr(preflight_mod, "EXCEL_PREFLIGHT_FULL_SCAN_MAX_ROWS", 1)
    monkeypatch.setattr(preflight_mod, "EXCEL_PREFLIGHT_PREFIX_ROWS_PER_SHEET", 2)
    assert run_preflight(path).estimated_tokens == ref_tokens


def test_preflight_excel_empty_sheet_plus_data(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    from hermes.ingestion.preflight import run_preflight

    path = tmp_path / "mix.xlsx"
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "MostlyEmpty"
    ws1 = wb.create_sheet("WithData")
    ws1.append(["x" * 40])
    wb.save(path)

    result = run_preflight(path)
    assert result.file_type == FileType.EXCEL
    assert result.estimated_tokens == 10
