"""Generate synthetic test fixtures for Hermes tests.

Run this script once to create small Excel and PDF files in tests/fixtures/.
Requires openpyxl and pymupdf to be installed.

Committed eval assets under ``tests/fixtures/eval/`` (manifests + goldens + copies of
``sample.xlsx`` / ``sample_text.pdf``) are static baselines for ``hermes eval``; this
script refreshes the binary copies from the generated sources so they stay in sync when
the generator changes.
"""

from __future__ import annotations

import shutil
from pathlib import Path

# Marker substring in boilerplate-only eval regions; regression tests key off this for mocked LLM.
BOILERPLATE_EVAL_MARKER = "BOILERPLATE_EVAL_NEGATIVE"


def _boilerplate_filler(min_chars: int = 9000) -> str:
    """Enough text that chunk merging splits vehicle content from boilerplate (~2048 token cap)."""
    unit = (
        "Terms and conditions. No vehicle slip data. "
        f"{BOILERPLATE_EVAL_MARKER} "
        "This section is boilerplate only. "
    )
    out: list[str] = []
    while len("".join(out)) < min_chars:
        out.append(unit)
    return "".join(out)[:min_chars]


def _sync_eval_fixture_copies(fixtures: Path) -> None:
    """Copy canonical sample inputs into ``fixtures/eval/`` for frozen eval paths."""
    eval_dir = fixtures / "eval"
    eval_dir.mkdir(parents=True, exist_ok=True)
    for name in ("sample.xlsx", "sample_text.pdf"):
        src = fixtures / name
        dst = eval_dir / name
        if src.is_file():
            shutil.copy2(src, dst)


def main() -> None:
    fixtures = Path(__file__).parent / "fixtures"
    fixtures.mkdir(exist_ok=True)

    _create_sample_excel(fixtures / "sample.xlsx")
    _create_sample_pdf_text(fixtures / "sample_text.pdf")
    _create_empty_excel(fixtures / "empty.xlsx")
    _sync_eval_fixture_copies(fixtures)

    print(f"Fixtures created in {fixtures}")


def _create_sample_excel(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    headers = [
        "Marca", "Descripcion", "Modelo", "Numero Serie",
        "Tipo", "Cobertura", "Suma Asegurada", "Deducible",
    ]
    ws.append(headers)
    ws.append([
        "Toyota", "Corolla SE", 2023, "JTDS4RCE1P0000001",
        "Sedan", "Amplia", 350000.00, "5%",
    ])
    ws.append([
        "Honda", "Civic EX", 2022, "2HGFC2F60NH000002",
        "Sedan", "Basica", 280000.00, "10%",
    ])
    ws.append([
        "Ford", "Ranger XLT", 2024, "1FTER4FH0PLA00003",
        "Pickup", "Amplia", 620000.00, "5%",
    ])
    ws.append([
        "Nissan", "Versa Advance", 2023, "3N1CN8DV7PL000004",
        "Sedan", "Limitada", 310000.00, "15%",
    ])
    ws.append([
        "Chevrolet", "Aveo LT", 2021, "3G1TB5CF1ML000005",
        "Sedan", "Basica", 220000.00, "10%",
    ])

    # Second sheet: boilerplate only (eval negative chunk). Large enough that chunk_pages
    # does not merge it with the Vehicles sheet (token budget in chunker._merge_segments).
    ws2 = wb.create_sheet("Boilerplate")
    ws2.append(["Text"])
    body = _boilerplate_filler()
    row_len = 400
    for i in range(0, len(body), row_len):
        ws2.append([body[i : i + row_len]])

    wb.save(str(path))


def _create_sample_pdf_text(path: Path) -> None:
    import pymupdf

    doc = pymupdf.open()

    # Page 1: all vehicle slips (positive chunk after merge of this page only).
    page = doc.new_page(width=612, height=792)
    text = (
        "SLIP DE FLOTILLA VEHICULAR\n\n"
        "Aseguradora: Seguros Atlas S.A.\n"
        "Poliza: POL-2024-001234\n\n"
        "Vehiculo 1:\n"
        "  Marca: Toyota\n"
        "  Descripcion: Corolla SE\n"
        "  Modelo: 2023\n"
        "  Numero de Serie: JTDS4RCE1P0000001\n"
        "  Tipo: Sedan\n"
        "  Cobertura: Amplia\n"
        "  Suma Asegurada: $350,000.00\n"
        "  Deducible: 5%\n\n"
        "Vehiculo 2:\n"
        "  Marca: Honda\n"
        "  Descripcion: Civic EX\n"
        "  Modelo: 2022\n"
        "  Numero de Serie: 2HGFC2F60NH000002\n"
        "  Tipo: Sedan\n"
        "  Cobertura: Basica\n"
        "  Suma Asegurada: $280,000.00\n"
        "  Deducible: 10%\n\n"
        "Vehiculo 3:\n"
        "  Marca: Ford\n"
        "  Descripcion: Ranger XLT\n"
        "  Modelo: 2024\n"
        "  Numero de Serie: 1FTER4FH0PLA00003\n"
        "  Tipo: Pickup\n"
        "  Cobertura: Amplia\n"
        "  Suma Asegurada: $620,000.00\n"
        "  Deducible: 5%\n"
    )
    page.insert_text((72, 72), text, fontsize=11)
    del page

    # Page 2: boilerplate only (negative chunk). Must stay within one segment (<= usable text
    # chunk tokens) or chunk_pages splits one page into multiple chunk_indices. Still large
    # enough that merged(page1) + merged(page2) exceeds the merge budget vs page 1 alone.
    page2 = doc.new_page(width=612, height=792)
    boilerplate = _boilerplate_filler(7800)
    cols = (72, 320)
    col_idx = 0
    x_base = cols[col_idx]
    y = 72
    line_h = 10
    step = 95
    for i in range(0, len(boilerplate), step):
        segment = boilerplate[i : i + step]
        page2.insert_text((x_base, y), segment, fontsize=7)
        y += line_h
        if y > 740:
            col_idx += 1
            if col_idx >= len(cols):
                break
            x_base = cols[col_idx]
            y = 72
    del page2

    doc.save(str(path))
    doc.close()


def _create_empty_excel(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    wb.save(str(path))


if __name__ == "__main__":
    main()
